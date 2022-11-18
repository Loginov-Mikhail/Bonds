import openpyxl
import pandas as pd

dfOut = pd.read_csv('output.csv', sep=';', encoding='cp1251')
with pd.ExcelWriter('output.xlsx', engine='openpyxl', date_format='DD.MM.YYYY') as writer:
    dfOut.to_excel(writer, sheet_name='Выбор облигаций', index=False, float_format='%.2f')
wb = openpyxl.load_workbook('output.xlsx')
ws = wb.active
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 53
ws['M1'] = 'Сумма инвестиций'
ws['N1'] = 100000
ws.column_dimensions['C'].number_format = openpyxl.styles.numbers.FORMAT_DATE_DDMMYY
for i in range(len(dfOut)):
    ws[f'C{i + 2}'].number_format = 'DD.MM.YYYY'
ws.column_dimensions['J'].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE_00
wb.save('output.xlsx')
wb.close()